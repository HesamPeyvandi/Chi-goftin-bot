"""
Telegram bot handlers.

- save_message: stores every incoming group message in the database.
- summarize_chat: builds the (unchanged) summarization prompt from stored
  history and delegates the actual summarization to ai_provider.
- Admin-only commands: /groups, /setpermanent, /removepermanent.
"""

import datetime
import io
import logging
import re

import telebot
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import ai_provider
import config
import database as db

logger = logging.getLogger(__name__)

# Reverses _format_message() so exported rows can show time/sender/text
# in separate spreadsheet columns instead of one raw line.
MESSAGE_LINE_PATTERN = re.compile(
    r"^\[ساعت (?P<time>\d{2}:\d{2}) \| (?P<sender>[^\]]+?)"
    r"(?: \(فوروارد از (?P<forward>[^)]+)\))?\]: (?P<text>.*)$",
    re.DOTALL,
)

TEHRAN_TZ = datetime.timezone(datetime.timedelta(hours=3, minutes=30))

DEFAULT_SUMMARY_MESSAGE_COUNT = 10
GROUP_CHAT_TYPES = ("group", "supergroup")


def _is_admin(user_id: int) -> bool:
    return config.ADMIN_USER_ID != 0 and user_id == config.ADMIN_USER_ID


def _format_message(message: telebot.types.Message) -> str:
    user_name = message.from_user.first_name
    text = message.text
    message_time = datetime.datetime.fromtimestamp(message.date, tz=TEHRAN_TZ).strftime("%H:%M")

    forward_source = ""
    is_forwarded = False

    if message.forward_from_chat:
        forward_source = message.forward_from_chat.title
        is_forwarded = True
    elif message.forward_from:
        forward_source = message.forward_from.first_name
        is_forwarded = True
    elif message.forward_sender_name:
        forward_source = message.forward_sender_name
        is_forwarded = True

    if is_forwarded:
        return f"[ساعت {message_time} | {user_name} (فوروارد از {forward_source})]: {text}"
    return f"[ساعت {message_time} | {user_name}]: {text}"


def _parse_message_line(line: str) -> tuple[str, str, str, str]:
    """Split a stored formatted line back into (time, sender, forwarded_from, text)."""
    match = MESSAGE_LINE_PATTERN.match(line)
    if not match:
        return "", "", "", line
    return (
        match.group("time") or "",
        match.group("sender") or "",
        match.group("forward") or "",
        match.group("text") or "",
    )


def _write_export_cell(sheet, row: int, column: int, value: str) -> None:
    """Write a cell for the export sheet, guarding against two rendering pitfalls:

    - A leading blank line in the message (common when a user hits Enter before
      typing) makes the cell LOOK empty in Excel, since Excel only shows the
      first line of a cell unless wrap-text is on. Stripping leading/trailing
      whitespace fixes this without altering the message's actual content.
    - openpyxl auto-detects any string starting with "=" as a formula. Since
      chat messages can start with "=" too, we force the cell's data type back
      to plain text so Excel never tries to evaluate it as a formula.
    """
    value = value.strip() if isinstance(value, str) else value
    cell = sheet.cell(row=row, column=column, value=value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def register_handlers(bot: telebot.TeleBot) -> None:
    @bot.message_handler(commands=["summarize"])
    def summarize_chat(message: telebot.types.Message):
        chat_id = message.chat.id

        if not db.has_messages(chat_id):
            bot.reply_to(message, "هنوز پیامی برای خلاصه کردن وجود نداره!")
            return

        try:
            command_parts = message.text.split()
            requested_count = (
                int(command_parts[1]) if len(command_parts) > 1 else DEFAULT_SUMMARY_MESSAGE_COUNT
            )
        except ValueError:
            requested_count = DEFAULT_SUMMARY_MESSAGE_COUNT

        messages_to_summarize = db.get_recent_messages(chat_id, requested_count)
        prompt = ai_provider.build_summary_prompt(messages_to_summarize)

        bot.reply_to(message, "در حال خوندن پیام‌ها... ⏳")

        try:
            summary_text = ai_provider.generate_summary(prompt)
            final_message = f"خلاصه ی پیام ها:\n\n{summary_text}"
            bot.reply_to(message, final_message)
        except Exception as error:
            bot.reply_to(message, "متاسفانه تو خلاصه‌سازی مشکلی پیش اومد.")
            logger.error("Summarization failed for chat %s: %s", chat_id, error)

    @bot.message_handler(commands=["groups"])
    def list_groups(message: telebot.types.Message):
        if not _is_admin(message.from_user.id):
            return

        groups = db.get_all_groups()
        if not groups:
            bot.reply_to(message, "هنوز هیچ گروهی ثبت نشده.")
            return

        lines = []
        for chat_id, title, is_permanent in groups:
            status = "دائمی" if is_permanent else "پیش‌فرض"
            lines.append(f"- {title} | {chat_id} | {status}")

        bot.reply_to(message, "\n".join(lines))

    @bot.message_handler(commands=["setpermanent"])
    def set_permanent(message: telebot.types.Message):
        if not _is_admin(message.from_user.id):
            return
        _handle_permanent_toggle(bot, message, is_permanent=True)

    @bot.message_handler(commands=["removepermanent"])
    def remove_permanent(message: telebot.types.Message):
        if not _is_admin(message.from_user.id):
            return
        _handle_permanent_toggle(bot, message, is_permanent=False)

    @bot.message_handler(commands=["export"])
    def export_messages(message: telebot.types.Message):
        if not _is_admin(message.from_user.id):
            return

        command_parts = message.text.split()
        if len(command_parts) < 2:
            bot.reply_to(
                message,
                "استفاده: /export <chat_id>\nبرای پیدا کردن chat_id هر گروه از /groups استفاده کن.",
            )
            return

        try:
            chat_id = int(command_parts[1])
        except ValueError:
            bot.reply_to(message, "chat_id باید یک عدد باشه.")
            return

        if not db.group_exists(chat_id):
            bot.reply_to(message, "این گروه هنوز شناخته نشده.")
            return

        all_messages = db.get_all_messages(chat_id)
        if not all_messages:
            bot.reply_to(message, "این گروه هیچ پیام ذخیره‌شده‌ای نداره.")
            return

        status_message = bot.reply_to(message, f"در حال ساخت فایل اکسل ({len(all_messages)} پیام)... ⏳")

        try:
            group_title = db.get_group_title(chat_id) or str(chat_id)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Messages"
            # The content is mostly Persian (RTL). Without this, Excel's
            # auto-detected text direction can misjudge cells that end in an
            # emoji (a weak/neutral-direction character), rendering the text
            # outside the visible cell area even though the data is intact.
            sheet.sheet_view.rightToLeft = True

            headers = ["ردیف", "ساعت", "فرستنده", "فوروارد از", "متن پیام"]
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_index, value=header)

            for offset, line in enumerate(all_messages, start=0):
                row_number = offset + 1
                excel_row = row_number + 1  # +1 to skip the header row
                time_str, sender, forwarded_from, text = _parse_message_line(line)
                row_values = [row_number, time_str, sender, forwarded_from, text]
                for col_index, value in enumerate(row_values, start=1):
                    _write_export_cell(sheet, excel_row, col_index, value)

            # Let the message column wrap so multi-line messages are fully visible,
            # and force RTL reading order explicitly (belt-and-suspenders on top of
            # the sheet-level setting above) so mixed Persian+emoji text can't be
            # miscalculated as LTR and pushed outside the visible cell.
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=5):
                for cell in row:
                    cell.alignment = cell.alignment.copy(
                        wrap_text=(cell.column == 5),
                        vertical="top",
                        horizontal="right",
                        readingOrder=2,
                    )

            column_widths = [6, 10, 22, 22, 80]
            for index, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width

            file_buffer = io.BytesIO()
            workbook.save(file_buffer)
            file_buffer.seek(0)

            safe_title = re.sub(r"[^\w\-]+", "_", group_title, flags=re.UNICODE).strip("_") or str(chat_id)
            file_name = f"export_{safe_title}_{chat_id}.xlsx"

            bot.send_document(
                message.chat.id,
                document=file_buffer,
                visible_file_name=file_name,
                reply_to_message_id=message.message_id,
                caption=f"اکسپورت {len(all_messages)} پیام از «{group_title}»",
            )
        except Exception as error:
            bot.reply_to(message, "متاسفانه تو ساخت فایل اکسپورت مشکلی پیش اومد.")
            logger.error("Export failed for chat %s: %s", chat_id, error)
        finally:
            try:
                bot.delete_message(status_message.chat.id, status_message.message_id)
            except Exception:
                pass

    @bot.message_handler(func=lambda message: True)
    def save_message(message: telebot.types.Message):
        if message.chat.type not in GROUP_CHAT_TYPES:
            return
        if not message.text:
            return

        chat_id = message.chat.id
        db.upsert_group(chat_id, message.chat.title)

        formatted_message = _format_message(message)
        db.add_message(chat_id, formatted_message)


def _handle_permanent_toggle(bot: telebot.TeleBot, message: telebot.types.Message, is_permanent: bool) -> None:
    command_parts = message.text.split()
    if len(command_parts) < 2:
        bot.reply_to(message, "استفاده: /setpermanent <chat_id> یا /removepermanent <chat_id>")
        return

    try:
        chat_id = int(command_parts[1])
    except ValueError:
        bot.reply_to(message, "chat_id باید یک عدد باشه.")
        return

    updated = db.set_group_permanent(chat_id, is_permanent)
    if not updated:
        bot.reply_to(message, "این گروه هنوز شناخته نشده.")
        return

    status = "دائمی" if is_permanent else "پیش‌فرض"
    bot.reply_to(message, f"وضعیت ذخیره‌سازی گروه {chat_id} به «{status}» تغییر کرد.")
